import datetime
from dataclasses import dataclass, field
import openpyxl
from XPathModel1 import XPathModel1
from XPathModel2 import XPathModel2
from XPathModel3 import XPathModel3


def get_row_value(row):
    val = row[1].value
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    if val is None:
        return ""
    return val


@dataclass
class DataModel:
    screenshot_location: str = ""

    xpath_model_1: XPathModel1 = None
    xpath_model_2: XPathModel2 = None
    xpath_model_3: XPathModel3 = None

    typ_dokumentu: dict = None
    wojewodztwo_1: dict = None
    wojewodztwo_2: dict = None
    rodzaj_zamierzenia_budowlanego: dict = None
    kategoria_obiektu: dict = None

    user_input_data_file: str = r"parametry_wyszukiwania.xlsx"

    def read_general_data(self):
        sheet = openpyxl.load_workbook(self.user_input_data_file)['dane ogólne']
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Ścieżka do zrzutów ekranu":
                self.screenshot_location = row[1].value
                break
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Rodzaj wyszukiwania":
                break

        self.typ_dokumentu = {}
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value == "Typ dokumentu (Rejestr Wniosków i Decyzji)":
                for j, row2 in enumerate(sheet.iter_rows(min_row=i+2)):
                    if row2[0].value != None and j != 0:
                        break
                    self.typ_dokumentu[row2[1].value] = j

        self.rodzaj_zamierzenia_budowlanego = {}
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value == "Rodzaj zamierzenia budowlanego":
                for j, row2 in enumerate(sheet.iter_rows(min_row=i+2)):
                    if row2[0].value != None and j != 0:
                        break
                    self.rodzaj_zamierzenia_budowlanego[row2[1].value] = j

        self.wojewodztwo_1 = {}
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value == "Województwo 1":
                for j, row2 in enumerate(sheet.iter_rows(min_row=i+2)):
                    if row2[0].value != None and j != 0:
                        break
                    self.wojewodztwo_1[row2[1].value] = j

        self.wojewodztwo_2 = {}
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value == "Województwo 2":
                for j, row2 in enumerate(sheet.iter_rows(min_row=i+2)):
                    if row2[0].value != None and j != 0:
                        break
                    self.wojewodztwo_2[row2[1].value] = j

        self.kategoria_obiektu = {}
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            if row[0].value == "Kategoria obiektu":
                for j, row2 in enumerate(sheet.iter_rows(min_row=i+2)):
                    if row2[0].value != None and j != 0:
                        break
                    self.kategoria_obiektu[row2[1].value] = j
    def read_xpath_model_3(self):
        sheet = openpyxl.load_workbook(self.user_input_data_file)['wybór danych do zapisu']
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Numer wniosku":
                self.xpath_model_3.szczegoly_wniosku_nr = get_row_value(row)
            if row[0].value == "Stan prawny":
                self.xpath_model_3.stan_prawny = get_row_value(row)
            if row[0].value == "Województwo":
                self.xpath_model_3.wojewodztwo = get_row_value(row)
            if row[0].value == "Kod pocztowy":
                self.xpath_model_3.kod_pocztowy = get_row_value(row)
            if row[0].value == "Poczta":
                self.xpath_model_3.poczta = get_row_value(row)
            if row[0].value == "Miejscowość":
                self.xpath_model_3.miejscowosc = get_row_value(row)
            if row[0].value == "Ulica":
                self.xpath_model_3.ulica = get_row_value(row)
            if row[0].value == "Numer domu":
                self.xpath_model_3.numer_domu = get_row_value(row)
            if row[0].value == "Numer lokalu":
                self.xpath_model_3.numer_lokalu = get_row_value(row)
            if row[0].value == "Nr działki ew.":
                self.xpath_model_3.nr_dzialki_ew = get_row_value(row)
            if row[0].value == "Obręb ew.":
                self.xpath_model_3.obreb_ew = get_row_value(row)
            if row[0].value == "Jednostka ew.":
                self.xpath_model_3.jednostka_ew = get_row_value(row)
            if row[0].value == "Rodzaj obiektu":
                self.xpath_model_3.rodzaj_obiektu = get_row_value(row)
            if row[0].value == "Kategoria obiektu":
                self.xpath_model_3.kategoria_obiektu = get_row_value(row)
            if row[0].value == "Kubatura budynku":
                self.xpath_model_3.kubatura_budynku = get_row_value(row)
            if row[0].value == "Nazwa zamierzenia budowlanego":
                self.xpath_model_3.nazwa_zamierzenia_budowlanego = get_row_value(row)
            if row[0].value == "Rodzaj zamierzenia budowlanego":
                self.xpath_model_3.rodzaj_zamierzenia_budowlanego = get_row_value(row)
            if row[0].value == "Nazwisko inwestora":
                self.xpath_model_3.nazwisko_inwestora = get_row_value(row)
            if row[0].value == "Imię inwestora":
                self.xpath_model_3.imie_inwestora = get_row_value(row)
            if row[0].value == "Nazwisko projektanta":
                self.xpath_model_3.nazwisko_pojektanta = get_row_value(row)
            if row[0].value == "Imię projektanta":
                self.xpath_model_3.imie_pojektanta = get_row_value(row)
            if row[0].value == "Numer uprawnień budowlanych projektanta":
                self.xpath_model_3.numer_uprawnien_budowlanych_pojektanta = get_row_value(row)
            if row[0].value == "Pozostali projektanci":
                self.xpath_model_3.pozostali_projektanci_pojektanta = get_row_value(row)
            if row[0].value == "Nazwa organu":
                self.xpath_model_3.nazwa_organu = get_row_value(row)
            if row[0].value == "Adres organu":
                self.xpath_model_3.adres_organu = get_row_value(row)
            if row[0].value == "Numer ewidencyjny wniosku nadawany w urzędzie":
                self.xpath_model_3.numer_ewidencyjny_wniosku_nadawany_w_urzedzie = get_row_value(row)
            if row[0].value == "Data wpływu wniosku do urzędu":
                self.xpath_model_3.data_wplywu_wniosku_do_urzedu = get_row_value(row)
            if row[0].value == "Data rejestracji w systemie komputerowym":
                self.xpath_model_3.data_rejestracji_w_systemie_komputerowym = get_row_value(row)
            if row[0].value == "Numer ewidencyjny decyzji nadawany w urzędzie":
                self.xpath_model_3.numer_ewidencyjny_decyzji_nadawany_w_urzedzie = get_row_value(row)
            if row[0].value == "Data wydania decyzji":
                self.xpath_model_3.data_wydania_decyzji = get_row_value(row)
            if row[0].value == "Decyzja":
                self.xpath_model_3.decyzja = get_row_value(row)
            if row[0].value == "Czy inwestor był wezwany do uzupełnienia braków formalnych (art. 64 § 2 kpa)":
                self.xpath_model_3.czy_inwestor_byl_wezwany_do_uzupelnienia_brakow_formalnych = get_row_value(row)
            if row[0].value == "Data wysłania wezwania do uzupełnienia braków formalnych":
                self.xpath_model_3.data_wyslania_wezwania_do_uzupelnienia_brakow_formalnych = get_row_value(row)
            if row[0].value == "Data uzupełnienia braków formalnych":
                self.xpath_model_3.data_uzupelnienia_brakow_formalnych = get_row_value(row)
            if row[0].value == "Wniosek wycofany przez inwestora":
                self.xpath_model_3.wniosek_wycofany_przez_inwestora = get_row_value(row)
            if row[0].value == "Data wycofania wniosku":
                self.xpath_model_3.data_wycofania_wniosku = get_row_value(row)
            if row[0].value == "Wniosek bez rozpoznania":
                self.xpath_model_3.wniosek_bez_rozpoznania = get_row_value(row)
            if row[0].value == "Wniosek przekazany zgodnie z właściwością":
                self.xpath_model_3.wniosek_przekazany_zgodnie_z_wlasciwoscia = get_row_value(row)
            if row[0].value == "Czy inwestor był wezwany do uzupełnienia dokumentacji":
                self.xpath_model_3.czy_inwestor_byl_wezwany_do_uzupelnienia_dokumentacji = get_row_value(row)
            if row[0].value == "Data wysłania postanowienia":
                self.xpath_model_3.data_wyslania_postanowienia = get_row_value(row)
            if row[0].value == "Data otrzymania uzupełnienia dokumentacji/upływ terminu na uzupełnienie":
                self.xpath_model_3.data_otrzymania_uzupelnienia_dokumentacji_uplyw_terminu_na_uzupelnienie = get_row_value(row)
            if row[0].value == "Liczba dni":
                self.xpath_model_3.liczba_dni = get_row_value(row)
            if row[0].value == "Czy wniosek wymagał uzgodnień z wojewódzkim konserwatorem zabytków":
                self.xpath_model_3.czy_wniosek_wymagal_uzgodnien_z_wojewodzkim_konserwatorem_zabytkow = get_row_value(row)
            if row[0].value == "Data wysłania dokumentów do konserwatora":
                self.xpath_model_3.data_wyslania_dokumentow_do_konserwatora = get_row_value(row)
            if row[0].value == "Data otrzymania uzgodnień z konserwatorem":
                self.xpath_model_3.data_otrzymania_uzgodnien_z_konserwatorem = get_row_value(row)
            if row[0].value == "Zawieszenie postępowania":
                self.xpath_model_3.zawieszenie_postepowania = get_row_value(row)
            if row[0].value == "Data zawieszenia postępowania":
                self.xpath_model_3.data_zawieszenia_postepowania = get_row_value(row)
            if row[0].value == "Data podjęcia postępowania":
                self.xpath_model_3.data_podjecia_postepowania = get_row_value(row)
            if row[0].value == "Czy zaistniały inne przyczyny wydłużenia ustawowego czasu wydania decyzji":
                self.xpath_model_3.czy_zaistnialy_inne_przyczyny_wydluzenia_ustawowego_czasu_wydania_decyzji = get_row_value(row)
            if row[0].value == "Położenie na mapie":
                self.xpath_model_3.polozenie_na_mapie = get_row_value(row)
    def read_xpath_model_1(self):
        sheet = openpyxl.load_workbook(self.user_input_data_file)['dane wyszukiwania']
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Url":
                self.xpath_model_1.url = get_row_value(row)
            if row[0].value == "Rodzaj wyszukiwania":
                self.xpath_model_1.rodzaj_wyszukiwania = get_row_value(row)
            if row[0].value == "Typ dokumentu":
                self.xpath_model_1.typ_dokumentu = get_row_value(row)
            if row[0].value == "Numer ewidencyjny wniosku":
                self.xpath_model_1.numer_ewidencyjny_wniosku = get_row_value(row)
            if row[0].value == "Numer ewidencyjny decyzji":
                self.xpath_model_1.numer_ewidencyjny_decyzji = get_row_value(row)
            if row[0].value == "Data złożenia wniosku / zgłoszenia (od)":
                self.xpath_model_1.data_zlozenia_wniosku_od = get_row_value(row)
            if row[0].value == "Data złożenia wniosku / zgłoszenia (do)":
                self.xpath_model_1.data_zlozenia_wniosku_do = get_row_value(row)
            if row[0].value == "Data wydania decyzji (od)":
                self.xpath_model_1.data_wydania_decyzji_od = get_row_value(row)
            if row[0].value == "Data wydania decyzji (do)":
                self.xpath_model_1.data_wydania_decyzji_do = get_row_value(row)
            if row[0].value == "Województwo (1)":
                self.xpath_model_1.wojewodztwo_1 = get_row_value(row)
            if row[0].value == "Organ administracji":
                self.xpath_model_1.organ_administracji = get_row_value(row)
            if row[0].value == "Województwo (2)":
                self.xpath_model_1.wojewodztwo_2 = get_row_value(row)
            if row[0].value == "Miejscowość":
                self.xpath_model_1.miejscowosc = get_row_value(row)
            if row[0].value == "Ulica":
                self.xpath_model_1.ulica = get_row_value(row)
            if row[0].value == "Numer domu":
                self.xpath_model_1.numer_domu = get_row_value(row)
            if row[0].value == "Nr działki ew.":
                self.xpath_model_1.nr_dzialki_ew = get_row_value(row)
            if row[0].value == "Obręb ew.":
                self.xpath_model_1.obreb_ew = get_row_value(row)
            if row[0].value == "Arkusz mapy":
                self.xpath_model_1.arkusz_mapy = get_row_value(row)
            if row[0].value == "Rodzaj zamierzenia budowlanego":
                self.xpath_model_1.rodzaj_zamierzenia_budowlanego = get_row_value(row)
            if row[0].value == "Kategoria obiektu":
                self.xpath_model_1.kategoria_obiektu = get_row_value(row)
            if row[0].value == "Nazwa zamierzenia":
                self.xpath_model_1.nazwa_zamierzenia = get_row_value(row)

    def save_search_result_1(self, list_XPathModel2: list[XPathModel2]):
        headers = [
            "Data wpływu",
            "Inwestor",
            "Nazwa zamierzenia",
            "Stan prawny",
            "Data wydania decyzji",
            "Url",
        ]

        workbook = openpyxl.load_workbook(self.user_input_data_file)
        sheet = workbook['wynik wyszukiwania']
        for i, header in enumerate(headers):
            sheet.cell(row=1, column=i+1).value = header
        for i, x2model in enumerate(list_XPathModel2):
            sheet.cell(row=i+2, column=1).value = x2model.data_wplywu
            sheet.cell(row=i+2, column=2).value = x2model.inwestor
            sheet.cell(row=i+2, column=3).value = x2model.nazwa_zamieszkania
            sheet.cell(row=i+2, column=4).value = x2model.stan_prawny
            sheet.cell(row=i+2, column=5).value = x2model.data_wydania_decyzji
            sheet.cell(row=i+2, column=6).value = x2model.url

        # for row in sheet.iter_rows(min_row=2):
        #     if row[0].value == "Ścieżka do zrzutów ekranu":
        #         self.screenshot_location = row[1].value
        #         break

        workbook.save(self.user_input_data_file)




    def __post_init__(self):
        self.read_general_data()
        self.xpath_model_1 = XPathModel1()
        self.xpath_model_2 = XPathModel2(0)
        self.xpath_model_3 = XPathModel3()
        self.read_xpath_model_1()
        self.read_xpath_model_3()
        a = 5

    def save_search_result_2(self, result):
        pass


if __name__ == "__main__":
    model = DataModel()
    a = 0
