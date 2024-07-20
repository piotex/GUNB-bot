import datetime
from dataclasses import dataclass, field
import openpyxl

from XPathModel1 import XPathModel1
from XPathModel2 import XPathModel2

from XPathModel3 import XPathModel3


def get_row_value(row):
    val = row[1].value
    if isinstance(val, datetime.date):
        return val.strftime('%Y-%m-%d')
    if val is None:
        return ""
    return val


@dataclass
class DataModel:
    screenshot_location: str = ""

    xpath_model_1: XPathModel1 = None
    xpath_model_2: XPathModel2 = None
    xpath_model_3: XPathModel3 = None

    user_input_data_file: str = r"parametry_wyszukiwania.xlsx"

    def read_general_data(self):
        sheet = openpyxl.load_workbook(self.user_input_data_file)['dane ogólne']
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Ścieżka do zrzutów ekranu":
                self.screenshot_location = row[1].value



    def read_xpath_model_1(self):
        sheet = openpyxl.load_workbook(self.user_input_data_file)['dane wyszukiwania']
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == "Url":
                self.xpath_model_1.url = get_row_value(row)
            if row[0].value == "Rodzaj wyszukiwania":
                self.xpath_model_1.rodzaj_wyszukiwania = get_row_value(row)
            if row[0].value == "Typ dokumentu":
                self.xpath_model_1.typ_dokumentu = get_row_value(row)
            if row[0].value == "Numer ewidencyjny wniosku":
                self.xpath_model_1.numer_ewidencyjny_wniosku = get_row_value(row)
            if row[0].value == "Numer ewidencyjny decyzji":
                self.xpath_model_1.numer_ewidencyjny_decyzji = get_row_value(row)
            if row[0].value == "Data złożenia wniosku / zgłoszenia (od)":
                self.xpath_model_1.data_zlozenia_wniosku_od = get_row_value(row)
            if row[0].value == "Data złożenia wniosku / zgłoszenia (do)":
                self.xpath_model_1.data_zlozenia_wniosku_do = get_row_value(row)
            if row[0].value == "Data wydania decyzji (od)":
                self.xpath_model_1.data_wydania_decyzji_od = get_row_value(row)
            if row[0].value == "Data wydania decyzji (do)":
                self.xpath_model_1.data_wydania_decyzji_do = get_row_value(row)
            if row[0].value == "Województwo (1)":
                self.xpath_model_1.wojewodztwo_1 = get_row_value(row)
            if row[0].value == "Organ administracji":
                self.xpath_model_1.organ_administracji = get_row_value(row)
            if row[0].value == "Województwo (2)":
                self.xpath_model_1.wojewodztwo_2 = get_row_value(row)
            if row[0].value == "Miejscowość":
                self.xpath_model_1.miejscowosc = get_row_value(row)
            if row[0].value == "Ulica":
                self.xpath_model_1.ulica = get_row_value(row)
            if row[0].value == "Numer domu":
                self.xpath_model_1.numer_domu = get_row_value(row)
            if row[0].value == "Nr działki ew.":
                self.xpath_model_1.nr_dzialki_ew = get_row_value(row)
            if row[0].value == "Obręb ew.":
                self.xpath_model_1.obreb_ew = get_row_value(row)
            if row[0].value == "Arkusz mapy":
                self.xpath_model_1.arkusz_mapy = get_row_value(row)
            if row[0].value == "Rodzaj zamierzenia budowlanego":
                self.xpath_model_1.rodzaj_zamierzenia_budowlanego = get_row_value(row)
            if row[0].value == "Kategoria obiektu":
                self.xpath_model_1.kategoria_obiektu = get_row_value(row)
            if row[0].value == "Nazwa zamierzenia":
                self.xpath_model_1.nazwa_zamierzenia = get_row_value(row)

    # def __init__(self):
    #     pass
    def __post_init__(self):
        self.read_general_data()
        self.xpath_model_1 = XPathModel1()
        self.read_xpath_model_1()
        a = 5


if __name__ == "__main__":
    model = DataModel()
    a = 0
